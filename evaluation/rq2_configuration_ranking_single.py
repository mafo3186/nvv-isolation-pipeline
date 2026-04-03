from __future__ import annotations

from pathlib import Path
from typing import Optional

import pandas as pd

from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

from evaluation.eval_io import write_csv_atomic
from utils.io import ensure_dir
from config.path_factory import (
    get_eval_mode_dir,
    get_summary_all_csv_path,
    get_global_combo_ranking_csv_path,
    get_global_evaluation_xlsx_path,
)


REQUIRED_COLUMNS = {
    "audio_id",
    "vad_mask",
    "asr_audio_in",
    "combo_key",
    "nvv_file",
    "mode",
    "n_gt",
    "n_cand",
    "tp",
    "fn",
    "fp",
    "recall",
    "precision",
    "f1",
    "mean_dice_eos_tp",
    "dice_eos_recall",
    "mean_overlap_s_tp",
    "insertion_rate",
    "deletion_rate",
    "error_rate",
}


NUMERIC_COLUMNS = [
    "n_gt",
    "n_cand",
    "tp",
    "fn",
    "fp",
    "precision",
    "recall",
    "f1",
    "mean_dice_eos_tp",
    "dice_eos_recall",
    "mean_overlap_s_tp",
    "insertion_rate",
    "deletion_rate",
    "error_rate",
]


def _read_summary_all_csv(summary_csv: Path) -> pd.DataFrame:
    """
    Read summary_all.csv with strict validation.

    Arguments:
        summary_csv: Path to summary_all.csv
    """
    if not summary_csv.exists():
        raise FileNotFoundError(f"Missing required file: {summary_csv}")

    try:
        df = pd.read_csv(summary_csv)
    except Exception as e:
        raise RuntimeError(f"Failed to read CSV: {summary_csv} ({e})") from e

    if df.shape[0] == 0:
        raise RuntimeError(f"summary_all.csv contains 0 rows: {summary_csv}")

    missing = sorted(REQUIRED_COLUMNS - set(df.columns))
    if missing:
        raise KeyError(f"summary_all.csv missing columns: {missing}")

    # Coerce numeric columns robustly
    for c in NUMERIC_COLUMNS:
        df[c] = pd.to_numeric(df[c], errors="coerce")

    # Normalize mode values
    df["mode"] = df["mode"].astype(str)

    return df


def compute_single_configuration_ranking(
    *,
    evaluation_dir: Path,
    mode: str,
    top_n: Optional[int] = None,
) -> pd.DataFrame:
    """
    Aggregate per-audio summary rows into a combo_key-level ranking.
    Writes out global_combo_ranking_{mode}.csv and returns the ranking DataFrame.

    Arguments:
        evaluation_dir: WorkspacePaths.evaluation (workspace/global/evaluation)
        mode: "full_gt" or "part_gt"
        top_n: Optional pre-filter after sorting
    """
    if mode not in ("full_gt", "part_gt"):
        raise ValueError(f"mode must be 'full_gt' or 'part_gt', got: {mode}")
    
    evaluation_dir = Path(evaluation_dir)
    mode_dir = get_eval_mode_dir(evaluation_dir, mode)
    ensure_dir(mode_dir)

    summary_csv = get_summary_all_csv_path(evaluation_dir, mode)
    df = _read_summary_all_csv(summary_csv)

    # Safety: keep only requested mode
    df = df[df["mode"] == mode].copy()
    if df.shape[0] == 0:
        raise RuntimeError(f"No rows found for mode={mode} in {summary_csv}")
    
    # --- Sanity checks for precision/f1 depending on mode ---
    if mode == "part_gt":
        # In part_gt, precision/f1 must be NaN (otherwise you accidentally computed full_gt metrics)
        if df["precision"].notna().any() or df["f1"].notna().any():
            bad = df[df["precision"].notna() | df["f1"].notna()][["audio_id", "combo_key", "precision", "f1"]].head(10)
            raise RuntimeError(f"part_gt contains non-NaN precision/f1 (showing first rows):\n{bad}")

    elif mode == "full_gt":
        # In full_gt, precision/f1 must be finite (NaN indicates broken computation / coercion)
        if df["precision"].isna().any() or df["f1"].isna().any():
            bad = df[df["precision"].isna() | df["f1"].isna()][["audio_id", "combo_key", "precision", "f1"]].head(10)
            raise RuntimeError(f"full_gt contains NaN precision/f1 (showing first rows):\n{bad}")

    else:
        raise ValueError(f"Unknown mode: {mode!r}")

    agg = (
        df.groupby(["combo_key", "vad_mask", "asr_audio_in"], dropna=False)
        .agg(
            n_audio_ids=("audio_id", "nunique"),
            macro_mean_n_gt=("n_gt", "mean"),
            macro_mean_n_cand=("n_cand", "mean"),
            macro_mean_tp=("tp", "mean"),
            macro_mean_fn=("fn", "mean"),
            macro_mean_fp=("fp", "mean"),
            macro_mean_recall=("recall", "mean"),
            macro_mean_precision=("precision", "mean"),
            macro_mean_f1=("f1", "mean"),
            macro_mean_mean_dice_eos_tp=("mean_dice_eos_tp", "mean"),
            macro_mean_dice_eos_recall=("dice_eos_recall", "mean"),
            macro_mean_mean_overlap_s_tp=("mean_overlap_s_tp", "mean"),
            macro_mean_insertion_rate=("insertion_rate", "mean"),
            macro_mean_deletion_rate=("deletion_rate", "mean"),
            macro_mean_error_rate=("error_rate", "mean"),
        )
        .reset_index()
    )

    agg.insert(0, "mode", mode)

    # Sorting rules per mode
    if mode == "full_gt":
        sort_cols = ["macro_mean_f1", "macro_mean_recall","macro_mean_dice_eos_recall", "macro_mean_mean_dice_eos_tp", "macro_mean_insertion_rate"]
        ascending = [False, False, False, False, True]
    else:
        sort_cols = ["macro_mean_recall", "macro_mean_dice_eos_recall", "macro_mean_mean_dice_eos_tp", "macro_mean_insertion_rate"]
        ascending = [False, False, False, True]

    agg = agg.sort_values(by=sort_cols, ascending=ascending, na_position="last").reset_index(drop=True)

    if top_n is not None:
        if int(top_n) <= 0:
            raise ValueError(f"top_n must be > 0, got: {top_n}")
        agg = agg.head(int(top_n)).copy()

    out_csv = get_global_combo_ranking_csv_path(evaluation_dir, mode)
    write_csv_atomic(agg, out_csv)

    return agg


def add_sheet_to_global_xlsx(
    *,
    evaluation_dir: Path,
    mode: str,
    sheet_name: str,
    df: pd.DataFrame,
) -> None:
    """
    Add or replace a sheet in the existing _global_evaluation_{mode}.xlsx.

    Arguments:
        evaluation_dir: WorkspacePaths.evaluation
        mode: "full_gt" or "part_gt"
        sheet_name: Sheet name to write (e.g., "global_combo_ranking")
        df: DataFrame to write
    """
    if mode not in ("full_gt", "part_gt"):
        raise ValueError(f"mode must be 'full_gt' or 'part_gt', got: {mode}")

    evaluation_dir = Path(evaluation_dir)

    xlsx_path = get_global_evaluation_xlsx_path(evaluation_dir, mode)
    if not xlsx_path.exists():
        raise FileNotFoundError(f"Missing workbook to extend: {xlsx_path}")

    tmp_path = xlsx_path.with_suffix(xlsx_path.suffix + ".tmp")

    wb = load_workbook(xlsx_path)

    if sheet_name in wb.sheetnames:
        ws_old = wb[sheet_name]
        wb.remove(ws_old)

    ws_new = wb.create_sheet(title=sheet_name)

    for row in dataframe_to_rows(df, index=False, header=True):
        ws_new.append(row)

    wb.save(tmp_path)
    tmp_path.replace(xlsx_path)


def run_single_configuration_ranking(
    *,
    evaluation_dir: Path,
    mode: str,
    top_n: Optional[int] = None,
    write_xlsx_sheet: bool = True,
) -> pd.DataFrame:
    """
    Convenience wrapper: compute ranking CSV and optionally add XLSX sheet.

    Arguments:
        evaluation_dir: WorkspacePaths.evaluation
        mode: "full_gt" or "part_gt"
        top_n: Optional cutoff
        write_xlsx_sheet: If True, add "global_combo_ranking" tab
    """
    ranking_df = compute_single_configuration_ranking(
        evaluation_dir=evaluation_dir,
        mode=mode,
        top_n=top_n,
    )

    if write_xlsx_sheet:
        add_sheet_to_global_xlsx(
            evaluation_dir=evaluation_dir,
            mode=mode,
            sheet_name=f"global_combo_ranking_{mode}",
            df=ranking_df,
        )

    return ranking_df