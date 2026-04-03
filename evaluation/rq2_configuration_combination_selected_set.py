#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Global Evaluation – Best-k Union Evaluation (artifact-based)
-----------------------------------------------------------

Reads selected best-k set (mode-scoped):
  - <evaluation_dir>/<mode>/global_f1_vs_k_<mode>.csv (preferred)
  OR
  - <evaluation_dir>/<mode>/global_best_k_set_<mode>.csv (fallback)

Then evaluates UNION of candidate events across selected combo_key tracks
for each audio_id (macro over audio_ids).

Writes (mode-scoped):
  - <evaluation_dir>/<mode>/global_best_k_union_per_audio_<mode>.csv
  - <evaluation_dir>/<mode>/global_best_k_union_summary_<mode>.csv
  - <evaluation_dir>/<mode>/global_best_k_union_set_<mode>.csv

Optionally appends sheets to:
  - <evaluation_dir>/<mode>/_global_evaluation_<mode>.xlsx
  (creates it if missing)

Notes:
- Strictly artifact-based (reads per_audio/.../annotations/nvv/*.json).
- Missing candidate files raise FileNotFoundError (no silent skipping).
"""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

from evaluation.eval_adapter_candidates import load_candidate_events_from_nvv_json
from evaluation.eval_io import write_csv_atomic
from evaluation.eval_union import (
    evaluate_union_for_audio_id,
    list_audio_ids_from_workspace,
)
from config.params import EVAL_T_COLLAR, EVAL_PERCENTAGE_OF_LENGTH
from config.path_factory import (
    get_eval_mode_dir,
    get_global_f1_vs_k_csv_path,
    get_global_best_k_set_csv_path,
    get_global_best_k_union_per_audio_csv_path,
    get_global_best_k_union_summary_csv_path,
    get_global_best_k_union_set_csv_path,
    get_global_evaluation_xlsx_path,
)


# --- Helpers ---

def _read_selected_set_from_f1_vs_k(
    *,
    evaluation_dir: Path,
    mode: str,
    k_override: Optional[int] = None,
) -> Tuple[int, List[str], pd.DataFrame]:
    """
    Returns: (best_k, selected_combo_keys_in_order, f1_vs_k_df)

    Reads:
      <evaluation_dir>/<mode>/global_f1_vs_k_<mode>.csv
    """
    evaluation_dir = Path(evaluation_dir)
    path = get_global_f1_vs_k_csv_path(evaluation_dir, mode)

    if not path.exists():
        raise FileNotFoundError(f"Missing required file: {path}")

    df = pd.read_csv(path)
    if df.shape[0] == 0:
        raise RuntimeError(f"{path.name} contains 0 rows: {path}")

    # Determine best_k
    if k_override is not None:
        best_k = int(k_override)
    elif "best_k" in df.columns and pd.notna(df["best_k"].iloc[0]):
        best_k = int(df["best_k"].iloc[0])
    else:
        # fallback: argmax macro_mean_f1 (only meaningful for full_gt)
        if "macro_mean_f1" not in df.columns:
            raise KeyError(f"{path.name} missing 'macro_mean_f1' and no best_k provided.")
        best_k = int(df.loc[df["macro_mean_f1"].astype(float).idxmax(), "k"])

    # Select row for that k
    row = df[df["k"].astype(int) == best_k]
    if row.shape[0] != 1:
        raise RuntimeError(f"Could not uniquely select row for k={best_k} in {path}")

    selected_json = str(row["selected_set_json"].iloc[0])

    try:
        selected = json.loads(selected_json)
    except Exception as e:
        raise ValueError(f"Could not json-parse selected_set_json for k={best_k} in {path.name}: {e}")

    if not isinstance(selected, list) or not all(isinstance(x, str) for x in selected):
        raise ValueError("selected_set_json must decode to list[str].")

    return best_k, selected, df


def _read_selected_set_from_best_k_set_csv(
    *,
    evaluation_dir: Path,
    mode: str,
) -> Tuple[int, List[str], pd.DataFrame]:
    """
    Fallback reader if global_f1_vs_k_<mode>.csv is not available.

    Reads:
      <evaluation_dir>/<mode>/global_best_k_set_<mode>.csv

    Returns: (best_k, selected_combo_keys_in_order, set_df)
    """
    evaluation_dir = Path(evaluation_dir)
    path = get_global_best_k_set_csv_path(evaluation_dir, mode)

    if not path.exists():
        raise FileNotFoundError(f"Missing required file: {path}")

    df = pd.read_csv(path)
    if df.shape[0] == 0:
        raise RuntimeError(f"{path.name} contains 0 rows: {path}")

    # preferred: selected_set_json
    if "selected_set_json" in df.columns and df["selected_set_json"].notna().any():
        selected_json = str(df["selected_set_json"].dropna().iloc[0])
        try:
            selected = json.loads(selected_json)
        except Exception as e:
            raise ValueError(f"Could not json-parse selected_set_json in {path.name}: {e}")
        if not isinstance(selected, list) or not all(isinstance(x, str) for x in selected):
            raise ValueError("selected_set_json must decode to list[str].")
        best_k = int(len(selected))
        return best_k, selected, df

    # fallback: combo_key column + rank_in_set ordering
    if "combo_key" not in df.columns:
        raise KeyError(f"{path.name} missing 'combo_key' and no 'selected_set_json' present.")

    if "rank_in_set" in df.columns:
        df = df.sort_values("rank_in_set")
    selected = df["combo_key"].astype(str).tolist()
    best_k = int(len(selected))
    return best_k, selected, df


def _write_xlsx_append_sheets_atomic(
    *,
    out_path: Path,
    sheets: Dict[str, pd.DataFrame],
) -> None:
    """
    Append/replace sheets in an .xlsx atomically.
    If file doesn't exist, create it.
    """
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = out_path.with_suffix(out_path.suffix + ".tmp")

    if out_path.exists():
        wb = load_workbook(out_path)
    else:
        from openpyxl import Workbook
        wb = Workbook()
        # remove default sheet
        if wb.sheetnames:
            wb.remove(wb[wb.sheetnames[0]])

    # Replace sheet if exists
    for sheet_name, df in sheets.items():
        if sheet_name in wb.sheetnames:
            ws_old = wb[sheet_name]
            wb.remove(ws_old)
        ws = wb.create_sheet(title=sheet_name)

        # Write dataframe (header + rows)
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
            ws.append(row)

    wb.save(tmp_path)
    tmp_path.replace(out_path)
    
# --- Public API ---

def run_best_k_union_evaluation(
    *,
    workspace_root: Path,
    evaluation_dir: Path,
    gt_dict: Dict[str, List[dict]],
    mode: str,
    evaluable_ids: Optional[List[str]] = None,
    k_override: Optional[int] = None,
    dedup_eps_s: Optional[float] = None,
    preload: bool = True,
    verbose_missing: bool = True,
    t_collar: float = EVAL_T_COLLAR,
    percentage_of_length: float = EVAL_PERCENTAGE_OF_LENGTH,
    evaluate_onset: bool = True,
    evaluate_offset: bool = True,
    match_labels: bool = False,
    write_xlsx_sheet: bool = True,
) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """
    Evaluate UNION(best-k tracks) per audio_id and aggregate.
Returns:
        per_audio_df: one row per audio_id
        summary_df: one-row macro summary (+ meta like best_k, dedup_eps_s)
        set_df: selected set in order (1 row per combo_key)
    Writes (mode-scoped):
      - global_best_k_union_per_audio_<mode>.csv
      - global_best_k_union_summary_<mode>.csv
      - global_best_k_union_set_<mode>.csv
      - optionally appends sheets to _global_evaluation_<mode>.xlsx
    """
    if mode not in ("full_gt", "part_gt"):
        raise ValueError(f"mode must be 'full_gt' or 'part_gt', got: {mode}")

    workspace_root = Path(workspace_root)
    evaluation_dir = Path(evaluation_dir)
    mode_dir = get_eval_mode_dir(evaluation_dir, mode)
    mode_dir.mkdir(parents=True, exist_ok=True)

    # NOTE: kept for interface compatibility; strict behavior now raises on missing files.
    _ = bool(verbose_missing)

    if evaluable_ids is None:
        evaluable_ids = list_audio_ids_from_workspace(workspace_root)
    evaluable_ids = list(evaluable_ids)

    # --- Read selected set in full_gt-mode and part_gt-mode) ---
    try:
        if mode == "full_gt":
            best_k, selected_combo_keys, _ = _read_selected_set_from_f1_vs_k(
                evaluation_dir=evaluation_dir,
                mode=mode,
                k_override=k_override,
            )
        else:
            best_k, selected_combo_keys, _ = _read_selected_set_from_best_k_set_csv(
                evaluation_dir=evaluation_dir,
                mode=mode,
            )
    except FileNotFoundError as e:
        raise FileNotFoundError(
            f"Could not read selected set for mode '{mode}': {e}\n"
            "Make sure to run the best k / selected-set construction for this mode first."
        )
    except Exception as e:
        raise RuntimeError(f"Error reading selected set for mode '{mode}': {e}")

    match_params = {
        "match_labels": bool(match_labels),
        "evaluate_onset": bool(evaluate_onset),
        "evaluate_offset": bool(evaluate_offset),
        "t_collar": float(t_collar),
        "percentage_of_length": float(percentage_of_length),
    }

    cache: Optional[Dict[Tuple[str, str], List[dict]]] = {} if preload else None

    per_rows: List[Dict[str, Any]] = []
    for aid in evaluable_ids:
        m = evaluate_union_for_audio_id(
            audio_id=aid,
            workspace_root=workspace_root,
            gt_dict=gt_dict,
            combo_keys_in_order=selected_combo_keys,
            load_candidate_events_fn=load_candidate_events_from_nvv_json,
            cache=cache,
            dedup_eps_s=dedup_eps_s,
            match_params=match_params,
            mode=mode,
        )
        per_rows.append(
            {
                "mode": mode,
                "audio_id": aid,
                "best_k": int(best_k),
                "dedup_eps_s": float(dedup_eps_s) if (dedup_eps_s is not None) else np.nan,
                "n_selected_tracks": int(len(selected_combo_keys)),
                "selected_set_json": json.dumps(selected_combo_keys),
                **m,
            }
        )

    per_audio_df = pd.DataFrame(per_rows)

    # Macro summary (NaN-aware: part_gt has NaN precision/f1 by design)
    def _nanmean(s: pd.Series) -> float:
        s2 = pd.to_numeric(s, errors="coerce")
        return float(s2.dropna().mean()) if s2.dropna().shape[0] > 0 else float("nan")

    macro_cols = [
        "n_gt", "n_cand", "tp", "fn", "fp",
        "insertion_rate", "deletion_rate", "error_rate",
        "precision", "recall", "f1",
        "mean_dice_eos_tp", "dice_eos_recall", "mean_overlap_s_tp",
    ]
    macro = {col: _nanmean(per_audio_df[col]) for col in macro_cols if col in per_audio_df.columns}

    summary_df = pd.DataFrame([{
        "mode": mode,
        "best_k": int(best_k),
        "n_audio_ids": int(len(evaluable_ids)),
        "dedup_eps_s": float(dedup_eps_s) if (dedup_eps_s is not None) else np.nan,
        "n_selected_tracks": int(len(selected_combo_keys)),
        "selected_set_json": json.dumps(selected_combo_keys),
        **{f"macro_mean_{k}": v for k, v in macro.items()},
    }])

    set_df = pd.DataFrame([{
        "mode": mode,
        "rank_in_set": i + 1,
        "combo_key": ck,
        "selected_set_json": json.dumps(selected_combo_keys),
        "best_k": int(best_k),
    } for i, ck in enumerate(selected_combo_keys)])

    # Write CSVs (mode-scoped)
    write_csv_atomic(per_audio_df, get_global_best_k_union_per_audio_csv_path(evaluation_dir, mode))
    write_csv_atomic(summary_df, get_global_best_k_union_summary_csv_path(evaluation_dir, mode))
    write_csv_atomic(set_df, get_global_best_k_union_set_csv_path(evaluation_dir, mode))

    # Optionally append to _global_evaluation_<mode>.xlsx (mode-scoped)
    if write_xlsx_sheet:
        xlsx_path = get_global_evaluation_xlsx_path(evaluation_dir, mode)
        sheets = {
            f"global_best_k_set_{mode}": set_df,
            f"global_best_k_union_summary_{mode}": summary_df,
            f"global_best_k_union_per_audio_{mode}": per_audio_df,
        }
        _write_xlsx_append_sheets_atomic(out_path=xlsx_path, sheets=sheets)

    return per_audio_df, summary_df, set_df